# Importer les modules requis
from flask import Flask, render_template, request, session,send_file, make_response
from flask_session import Session
import pandas as pd
import numpy as np
import os
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from unidecode import unidecode
import tkinter as tk
from tkinter import filedialog
from datetime import datetime

# Initialiser l'application Flask
app = Flask(__name__)

# Configurer la session Flask
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

# Variable pour stocker le nombre choisi
day = None
df_productivity = None
df_hubplanner = None
df_productivity1 = None
df_timesheet = None
df_table_final_s = None
df_table_final_suu1 = None

# Définir la route principale de l'application pour la méthode GET
@app.route('/', methods=['GET', 'POST'])
def index():
    global day

    if request.method == 'POST':
        # Récupérer le nombre sélectionné depuis le formulaire
        day = int(request.form['number'])
    return render_template('index.html')

# Définir les routes pour charger les fichiers
@app.route('/load_productivity', methods=['POST'])
def load_productivity():
    global df_productivity
    file = request.files['productivity_file']
    df_productivity = pd.read_excel(file)
    session["df_productivity"] = df_productivity.to_dict()
    return df_productivity.head().to_html()



@app.route('/load_productivity1', methods=['POST'])
def load_productivity1():
    global df_productivity1
    file = request.files['productivity1_file']
    df_productivity1 = pd.read_excel(file)
    session["df_productivity1"] = df_productivity1.to_dict()
    return df_productivity1.head().to_html()



@app.route('/load_hubplanner', methods=['POST'])
def load_hubplanner():
    global df_hubplanner
    file = request.files['hubplanner_file']
    df_hubplanner = pd.read_excel(file)
    session["df_hubplanner"] = df_hubplanner.to_dict()
    return df_hubplanner.head().to_html()



@app.route('/load_timesheet', methods=['POST'])
def load_timesheet():
    global df_timesheet
    file = request.files['timesheet_file']
    df_timesheet = pd.read_excel(file)
    session["df_timesheet"] = df_timesheet.to_dict()
    return df_timesheet.head().to_html()



## Ajouter le deuxième code à l'intérieur de la route compare_files
# Définir la route pour comparer les fichiers
@app.route('/compare_files', methods=['POST'])
def compare_files():
    global df_productivity, df_hubplanner, df_table_final_s
    
    # Récupérer les DataFrames depuis la session
    df_productivity = pd.DataFrame.from_dict(session.get("df_productivity", {}))
    df_hubplanner = pd.DataFrame.from_dict(session.get("df_hubplanner", {}))

    # Vérifier que les deux DataFrames ont été chargés
    if df_productivity.empty or df_hubplanner.empty:
        return "Veuillez d'abord charger les deux fichiers."

    

    start_date = request.form['start_date']
    end_date = request.form['end_date']

    # Vérifier si les dates ont été saisies
    if  start_date or end_date:
        start_day = int(start_date.split('-')[2])  # Récupérer le jour du mois de la date de début
        end_day = int(end_date.split('-')[2])

        df_productivity = df_productivity[df_productivity['Date'].dt.day <= end_day]
        df_productivity = df_productivity[df_productivity['Date'].dt.day >= start_day]
        

        


    ##PRODUCTIVITY
    pivot_tab_prod = pd.pivot_table(df_productivity, index=["TeamMember", "Project", "PRODUCTIVITY_Id"], values=["MD"], aggfunc=np.sum, fill_value=0)
    pivot_tab_prod_sorted_prodvt = pivot_tab_prod.sort_values(by=["TeamMember"])
    print("suuu")
    print("Productivity DataFrame:")
    print(pivot_tab_prod_sorted_prodvt)

    ##HUB PLANNER
    df_hb_table = df_hubplanner[["Resource/Unassigned Name", "Name", "Date Range (scheduled)", "Project/Event Code", "First Name", "Last Name"]]

    # Trier le dataframe par la colonne "Resource/Unassigned Name"
    df_hb_table.sort_values(by="Resource/Unassigned Name", inplace=True)

    # Réinitialiser les index après le tri
    df_hb_table.reset_index(drop=True, inplace=True)

    print("HubPlanner DataFrame:")
    print(df_hb_table)

    noms_colonnes_table = ['Name','Project', 'MD_Productivity', 'MD_HubPlanner', 'Gap']
    print(noms_colonnes_table)

    # Créez un DataFrame vide avec les noms de colonnes
    df_table_final = pd.DataFrame(columns=noms_colonnes_table)
    # Copier le contenu du DataFrame original dans la copie
    df_copie_prty = pivot_tab_prod_sorted_prodvt.copy()
    df_copie_hupplanner = df_hb_table.copy()

    for team_member_value, row in pivot_tab_prod_sorted_prodvt.iterrows():
        productivity_id_value = team_member_value[2]
        md_value = row["MD"]
        print(team_member_value[0],"name11111")

        for _,row in df_hb_table.iterrows():
            productivity_id_value_hb = row["Project/Event Code"]
            resource_name=row["Resource/Unassigned Name"]
            projectname = row["Name"]
            md_hubplanner=row["Date Range (scheduled)"]
            #print("test")
            firstname=row["First Name"]
            lastname=row["Last Name"]

            concat=str(lastname) + " " + firstname

            name1=unidecode(str(team_member_value[0]).lower())
            name2=unidecode(str(concat).lower())
                       
            if name1.replace(' ', '') == name2.replace(' ', '') and productivity_id_value==productivity_id_value_hb :
                print("cool")
                
                gap=float(md_hubplanner) - float(md_value)
                nouvelle_ligne = pd.DataFrame([[name1,projectname,md_value, md_hubplanner, gap]], columns =['Name','Project', 'MD_Productivity', 'MD_HubPlanner', 'Gap'])
                df_table_final = pd.concat([df_table_final, nouvelle_ligne], ignore_index=True)

                df_copie_hupplanner = df_copie_hupplanner.drop(df_copie_hupplanner[(df_copie_hupplanner['Resource/Unassigned Name'] == resource_name ) & (df_copie_hupplanner['Project/Event Code'] == productivity_id_value_hb)].index)

                filtre = (df_copie_prty.index.get_level_values('TeamMember') == team_member_value[0]) & \
                (df_copie_prty.index.get_level_values('Project') == team_member_value[1]) & \
                (df_copie_prty.index.get_level_values('PRODUCTIVITY_Id') == team_member_value[2])
                df_copie_prty = df_copie_prty.drop(df_copie_prty[filtre].index)
                break


    print(df_table_final)
    for team_member_value, row in df_copie_prty.iterrows():
        name1=str(team_member_value[0].lower())
        md_value = row["MD"]

        for _,row2 in df_table_final.iterrows():
            name2=str(row2["Name"]).lower()

            if unidecode(name1.replace(' ', '')) == unidecode(name2.replace(' ', '')) :
                print("cool")
                nouvelle_ligne = pd.DataFrame([[name1,team_member_value[1],md_value, 0, md_value]], columns =['Name','Project', 'MD_Productivity', 'MD_HubPlanner', 'Gap'])
                df_table_final = pd.concat([df_table_final, nouvelle_ligne], ignore_index=True)
                break

    
    print(df_table_final)
    for _,row in df_copie_hupplanner.iterrows():
        resource_name=row["Resource/Unassigned Name"]
        projectname = row["Name"]
        md_hubplanner=row["Date Range (scheduled)"]
        firstname=row["First Name"]
        lastname=row["Last Name"]

        concat=str(lastname) + " " + firstname
        name1=str(concat).lower()

        for _,row2 in df_table_final.iterrows():
            name2=str(row2["Name"]).lower()

            if unidecode(name1.replace(' ', '')) == unidecode(name2.replace(' ', '')) and projectname != "TRA" and projectname != "OOO" and projectname != "Vacation" :
                print("cool")
                nouvelle_ligne = pd.DataFrame([[name1,projectname, 0, md_hubplanner, md_hubplanner]], columns =['Name','Project', 'MD_Productivity', 'MD_HubPlanner', 'Gap'])
                df_table_final = pd.concat([df_table_final, nouvelle_ligne], ignore_index=True)
                break




    df_table_final_s = pd.DataFrame(columns=noms_colonnes_table)
    # Grouper par la colonne "Name"
    groupes = df_table_final.groupby('Name')
    print(df_table_final)
    # Parcourir chaque groupe
    for name, groupe in groupes:
        # Calculer la somme des colonnes numériques pour le groupe
        somme_groupes = groupe.select_dtypes(include='number').sum()
        nouvelle_ligne = pd.DataFrame([[name, '', somme_groupes['MD_Productivity'], somme_groupes['MD_HubPlanner'], somme_groupes['MD_Productivity']-somme_groupes['MD_HubPlanner']]], columns=df_table_final.columns)
        df_table_final_s = pd.concat([nouvelle_ligne, groupe, df_table_final_s], ignore_index=True)

    print(df_table_final_s)
    tampon_ligne=df_table_final_s.loc[0, 'Name']
    taille_max = df_table_final_s.shape[0]
    i=1
    # Boucle while pour itérer sur le DataFrame
    while i < taille_max:
        # Accéder à la ligne spécifique avec l'index "i"
        ligne_i = df_table_final_s.loc[i, 'Name']
        if unidecode(ligne_i.replace(' ', '')) == unidecode(tampon_ligne.replace(' ', '')): 
            df_table_final_s.iloc[i, df_table_final_s.columns.get_loc('Name')] = ''
        tampon_ligne = ligne_i

        i += 1

      # Utiliser tempfile pour créer un chemin d'accès temporaire pour le fichier Excel
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        gap_analysis = temp_file.name

        # Sauvegarder le DataFrame en tant que fichier Excel
        df_table_final_s.to_excel(gap_analysis, index=False)

        # Charger le fichier Excel avec openpyxl
        wb = load_workbook(gap_analysis)

        # Sélectionner la feuille de calcul
        ws = wb.active

        # Appliquer un style aux cellules de la colonne "gap" égales à 0
        for cell in ws["E"]:
            if cell.value == 0:
                cell.fill = PatternFill(start_color="9FFF9F", end_color="9FFF9F", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFABAB", end_color="FFABAB", fill_type="solid")

        # Enregistrer les modifications dans le fichier Excel
        wb.save(gap_analysis)

    # Envoyer le fichier Excel au client pour téléchargement avec le bon nom de fichier
    response = make_response(send_file(gap_analysis, as_attachment=True))
    response.headers['Content-Disposition'] = 'attachment; filename=gap_analysis.xlsx'
    return response






@app.route('/compare_timesheet_files', methods=['POST'])
def compare_timesheet_files():  # Correction: Renommer la fonction pour la rendre unique
    global df_productivity1, df_timesheet, df_table_final_suu1
    
    # Récupérer les DataFrames depuis la session
    df_productivity1 = pd.DataFrame.from_dict(session.get("df_productivity1", {}))
    df_timesheet = pd.DataFrame.from_dict(session.get("df_timesheet", {}))

    # Vérifier que les deux DataFrames ont été chargés
    if df_productivity1.empty or df_timesheet.empty:
        return "Veuillez d'abord charger les deux fichiers."

    print("productivity1")
    print(df_productivity1)
    print("timesheet TYPEEEEE")
    print(df_timesheet['Month'])


    month_name_to_number = {
        'january': 1,
        'february': 2,
        'march': 3,
        'april': 4,
        'may': 5,
        'june': 6,
        'july': 7,
        'august': 8,
        'september': 9,
        'october': 10,
        'november': 11,
        'december': 12
    }


    selected_month_name = request.form.get('selected_month')
    selected_month_number = month_name_to_number.get(selected_month_name)

    print("Month sélectionné : ", selected_month_number, " name " ,selected_month_name)

    # Filtrer le DataFrame df_timesheet en fonction du mois sélectionné
    #filtered_df_timesheet = df_timesheet[df_timesheet['Month'] == selected_month_number]
    num_mois = int(selected_month_number)
    
    #filtered_df_timesheet = df_timesheet[df_timesheet['Month'] == 5]
    #filtered_df_timesheet = df_timesheet[df_timesheet['Month'] == 5]


    print("Valeur de num_mois :", num_mois)
    filtered_df_timesheet = df_timesheet[df_timesheet['Month'] == num_mois]
    print("Nombre de lignes après le filtrage :", len(filtered_df_timesheet))

    print("#############################################")
    print(filtered_df_timesheet)
    col_type = df_timesheet["Month"].dtype
    print("Le type de la colonne 'Month' est :", col_type)
    print("#############################################")
    

    pivot_tab_prod = pd.pivot_table(df_productivity1, index=["TeamMember"], values=["MD"], aggfunc=np.sum, fill_value=0)

    # Filtrer la table pivot en fonction de la liste des membres d'équipe
    team_members_list = pivot_tab_prod.index.tolist()
    filtered_pivot_tab_prod2 = pivot_tab_prod.loc[pivot_tab_prod.index.isin(team_members_list)]
    
    print("#############################################")
    

    

    pivot_table2 = pd.pivot_table(filtered_df_timesheet, index=["Employee"], values=["NbEvents"], columns=["EventType"], aggfunc=np.sum)
    print(pivot_table2)
    # Sélectionner les colonnes spécifiques (P1, P2, P3) et calculer la somme
    team_members_list2 = pivot_table2.index.tolist()

    filtered_pivot_table = pivot_table2.loc[pivot_table2.index.isin(team_members_list2)] 
    #TOUTES LES PERSONNES
    columns_to_sum = ['P1', 'P2', 'P3']
    sum_filtered_pivot_table = filtered_pivot_table.loc[:, filtered_pivot_table.columns.get_level_values(1).isin(columns_to_sum)].sum(axis=1)

    Timesheet_table = sum_filtered_pivot_table.to_frame()
    print("Tableau croiser dynamique de Timesheet")
    print(Timesheet_table)

    new_name=['MD']
    filtered_pivot_tab_prod2.columns=new_name
    Timesheet_table.columns=new_name

    print("MD")
    print(filtered_pivot_tab_prod2)
    df_transpose = filtered_pivot_tab_prod2.transpose()
    df_transpose2 = Timesheet_table.transpose()
    print(" productivityyyyyy tranpose",df_transpose )
    print(" timesheeeeet tranpose",df_transpose2)

    liste_colonnes = df_transpose.columns.tolist()
    liste_colonnes_tampon = liste_colonnes


    liste_colonnes2 = df_transpose2.columns.tolist()
    liste_colonnes_tampon2 = liste_colonnes2
    print("liste 1",liste_colonnes)
    print("liste 2",liste_colonnes2)

    noms_colonnes_table = ['Name', 'MD_Productivity', 'MD_Timesheet', 'Gap']

    # Créez un DataFrame vide avec les noms de colonnes
    df_table_final_suu1 = pd.DataFrame(columns=noms_colonnes_table)

    nouvelle_liste_colonnes_tampon = []
    nouvelle_liste_colonnes_tampon2 = []

    for nom_colonne in liste_colonnes_tampon:
        for nom_colonne2 in liste_colonnes_tampon2:
            if str(nom_colonne.replace(" ", "")) == str(nom_colonne2.replace(" ", "")):
                print("meme nom")
                nouvelle_ligne = pd.DataFrame([[nom_colonne,float(df_transpose[nom_colonne].values), float(df_transpose2[nom_colonne2].values), float(df_transpose[nom_colonne].values) - float(df_transpose2[nom_colonne2].values)]], columns =['Name','MD_Productivity','MD_Timesheet','Gap'])
                df_table_final_suu1 = pd.concat([df_table_final_suu1, nouvelle_ligne], ignore_index=True)
                break
        else:
            nouvelle_liste_colonnes_tampon.append(nom_colonne)

    liste_colonnes_tampon = nouvelle_liste_colonnes_tampon
    liste_colonnes_tampon2 = nouvelle_liste_colonnes_tampon2

    for nom_colonne in liste_colonnes_tampon:
        nouvelle_ligne = pd.DataFrame([[nom_colonne,float(df_transpose[nom_colonne].values), 0, float(df_transpose[nom_colonne].values)]], columns =['Name','MD_Productivity','MD_Timesheet','Gap'])
        df_table_final_suu1 = pd.concat([df_table_final_suu1, nouvelle_ligne], ignore_index=True)


    for nom_colonne2 in liste_colonnes_tampon2:
        nouvelle_ligne2 = pd.DataFrame([[nom_colonne2,0, float(df_transpose2[nom_colonne2].values) , float(df_transpose2[nom_colonne2].values)]], columns =['Name','MD_Productivity','MD_Timesheet','Gap'])
        df_table_final_suu1 = pd.concat([df_table_final_suu1, nouvelle_ligne2], ignore_index=True)

    print(df_table_final_suu1)

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        gap_analysis_timesheet = temp_file.name

        # Sauvegarder le DataFrame en tant que fichier Excel
        df_table_final_suu1.to_excel(gap_analysis_timesheet, index=False)

        # Charger le fichier Excel avec openpyxl
        wb = load_workbook(gap_analysis_timesheet)

        # Sélectionner la feuille de calcul
        ws = wb.active

        # Appliquer un style aux cellules de la colonne "gap" égales à 0
        for cell in ws["D"]:
            if cell.value == 0:
                cell.fill = PatternFill(start_color="9FFF9F", end_color="9FFF9F", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFABAB", end_color="FFABAB", fill_type="solid")

        # Enregistrer les modifications dans le fichier Excel
        wb.save(gap_analysis_timesheet)

    # Envoyer le fichier Excel au client pour téléchargement avec le bon nom de fichier
    response = make_response(send_file(gap_analysis_timesheet, as_attachment=True))
    response.headers['Content-Disposition'] = 'attachment; filename=gap_analysis_timesheet.xlsx'
    return response









# Lancer l'application Flask
if __name__ == '__main__':
    app.run(debug=True)
